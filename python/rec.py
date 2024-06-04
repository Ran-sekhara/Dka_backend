import sys
from openpyxl import load_workbook

def add_recommendation(typerecommendation, rating, recommendation, age, gender, diabetesType, isSmoke, area):
    file_path = 'C:\\Users\\DELL\\Downloads\\dataset.xlsx'
    
    # Load the workbook and select the active worksheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    
    # Find the row where age, gender, diabetesType, isSmoke, and area match
    matching_row = None
    for row in range(2, sheet.max_row + 1):
        cell_age = sheet.cell(row=row, column=3).value
        cell_gender = sheet.cell(row=row, column=2).value
        cell_diabetesType = sheet.cell(row=row, column=4).value
        cell_isSmoke = sheet.cell(row=row, column=5).value
        cell_area = sheet.cell(row=row, column=6).value
        
        if (
            str(cell_age).strip().lower() == str(age).strip().lower() and
            str(cell_gender).strip().lower() == str(gender).strip().lower() and
            str(cell_diabetesType).strip().lower() == str(diabetesType).strip().lower() and
            str(cell_isSmoke).strip().lower() == str(isSmoke).strip().lower() and
            str(cell_area).strip().lower() == str(area).strip().lower()
        ):
            matching_row = row
            break
    
    if matching_row is not None:
        # Write recommendation and rating to the matching row
        column_mapping = {
            'Nutrition and Diet': 7,
            'Routine Physical Activities': 9,
            'Self-care': 11,
            'Psycho-social Care': 13
        }
        
        if typerecommendation in column_mapping:
            column = column_mapping[typerecommendation]
            
            # Retrieve the current value in the cell
            current_recommendation = sheet.cell(row=matching_row, column=column).value
            current_rating = sheet.cell(row=matching_row, column=column + 1).value

            # Concatenate the new recommendation with the existing one
            if current_recommendation:
                new_recommendation = f"{current_recommendation}; {recommendation}"
            else:
                new_recommendation = recommendation

            # Concatenate the new rating with the existing one
            if current_rating:
                new_rating = f"{current_rating}; {rating}"
            else:
                new_rating = rating

            # Write the updated recommendation and rating back to the cell
            sheet.cell(row=matching_row, column=column).value = new_recommendation
            sheet.cell(row=matching_row, column=column + 1).value = new_rating

            workbook.save(filename=file_path)
            print("Recommendation added successfully")
        else:
            print(f"Invalid category: {typerecommendation}")
    else:
        print("No matching row found")

if __name__ == '__main__':
    if len(sys.argv) != 9:
        print("Invalid number of arguments. Expected 8 arguments: typerecommendation, rating, recommendation, age, gender, diabetesType, isSmoke, area")
    else:
        typerecommendation = sys.argv[1]
        rating = sys.argv[2]
        recommendation = sys.argv[3]
        age = sys.argv[4]
        gender = sys.argv[5]
        diabetesType = sys.argv[6]
        isSmoke = sys.argv[7]
        area = sys.argv[8]
        add_recommendation(typerecommendation, rating, recommendation, age, gender, diabetesType, isSmoke, area)